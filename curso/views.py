import re
from django.shortcuts import redirect, get_object_or_404
from .models import TipoDocumento, Rol, Usuario,Programa, Departamento, Municipio, Curso, Solicitud, Aspirante
from .forms import UsuarioEditForm, UsuarioCreateForm, InicioSesionForm, CursoForm
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse, JsonResponse, FileResponse
from .forms import CursoForm, AspiranteForm
import os
import io
from django.conf import settings
from django.shortcuts import render
from docxtpl import DocxTemplate
from django.db.models import Q, OuterRef, Subquery # Importar Q para búsquedas complejas
import pandas as pd
from io import BytesIO
import zipfile
from datetime import datetime, timedelta
from openpyxl import load_workbook
from django.core.files.images import get_image_dimensions
from django.core.exceptions import ValidationError
from django.core.files.storage import FileSystemStorage
from django.views.decorators.csrf import csrf_exempt
from pdf2image import convert_from_path
import pytesseract
from PIL import Image
from django.db.models import Count
from PyPDF2 import PdfMerger   
import tempfile
import shutil

def dashboard(request):
    return render(request, 'dashboard.html')

#Buscar Cursos
def buscar_curso(request):
    q = request.GET.get("q", "").strip()
    courses = []

    if q:
        courses = Curso.objects.annotate(
            inscritos_count=Count("aprendices")
        ).filter(
            Q(programa__nombre__icontains=q) |
            Q(usuario__first_name__icontains=q) |
            Q(usuario__last_name__icontains=q) |
            Q(usuario__documento__icontains=q)
        )

        # Obtener el estado REAL de cada curso desde su solicitud
        for c in courses:
            try:
                # Buscar la solicitud asociada a este curso
                solicitud = Solicitud.objects.get(curso=c)
                estado_map = {0: "pending", 1: "approved", 2: "rejected"}
                c.status = estado_map.get(solicitud.estado, "pending")
            except Solicitud.DoesNotExist:
                # Si no existe solicitud, está pendiente
                c.status = "pending"
            
            c.registrationLink = c.link

    return render(request, "buscar_curso.html", {
        "courses": courses,
        "q": q
    })

#Eliminar Curso
def eliminar_curso(request, curso_id):
    if request.method == "POST":
        curso = get_object_or_404(Curso, id=curso_id)
        curso.delete()
        messages.success(request, f"El curso {curso.programa.nombre} fue eliminado correctamente.")
    return redirect("buscar_curso")

#Crear Solicitud
def crear_solicitud(request, curso_id):
    if request.method == "POST":
        curso = get_object_or_404(Curso, id=curso_id)

        # Buscar si ya hay una solicitud de este curso
        solicitud, creada = Solicitud.objects.get_or_create(
            curso=curso,
            defaults={"estado": 0}  # si no existe, se crea con estado=0
        )

        if not creada:
            # si ya existía, la actualizamos
            solicitud.estado = 0
            solicitud.save()

    return redirect("buscar_curso")  # 👈 cambia al nombre real de tu vista de búsqueda
    
#Coordinador
def coordinador(request):
    solicitudes = Solicitud.objects.select_related("curso__programa", "curso__usuario")

    # Creamos un diccionario de traducción
    estado_map = {
        0: "pending",
        1: "approved",
        2: "rejected",
    }

    # Agregamos un atributo extra a cada objeto
    for s in solicitudes:
        s.status = estado_map.get(s.estado, "unknown")

    pending_courses = [s for s in solicitudes if s.estado == 0]
    approved_courses = [s for s in solicitudes if s.estado == 1]
    rejected_courses = [s for s in solicitudes if s.estado == 2]
    return render(request, 'coordinador_dashboard.html',
                  { "pending_courses": pending_courses,
                    "approved_courses": approved_courses,
                    "rejected_courses": rejected_courses,
                    "courses": solicitudes,
                    })

#Aprobar-Rechazar Solicitudes
def approve_request(request, pk):
    solicitud = get_object_or_404(Solicitud, pk=pk)
    solicitud.estado = 1  # aprobado
    solicitud.save()
    return redirect("coordinador")  

def reject_request(request, pk):
    solicitud = get_object_or_404(Solicitud, pk=pk)
    solicitud.estado = 2  # rechazado
    solicitud.save()
    return redirect("coordinador")

# Ver detalles 
def curso_detalle(request, pk):
    curso = get_object_or_404(Curso, pk=pk)

    form = CursoForm(initial={
        # PROGRAMA (FK)
        "nombreprograma": curso.programa,  
        "codigoprograma": curso.programa.codigo,   
        "versionprograma": curso.programa.version, 
        "duracionprograma": curso.programa.duracion,

        # FECHAS
        "fechainicio": curso.fecha_inicio,
        "fechafin": curso.fecha_fin,

        # RESPONSABLE (usuario)
        "nombre": f"{curso.usuario.first_name} {curso.usuario.last_name}".strip(),
        "tipodoc": getattr(curso.usuario.tipo_documento, "nombre", ""),
        "numerodoc": getattr(curso.usuario, "documento", ""),
        "correo": curso.usuario.email,

        # EMPRESA / CARTA
        "empresa": curso.caracterizacion,   
        "carta_empresa": curso.carta,

        # ARCHIVOS / DOCUMENTOS
        "fecha1": curso.pdf_documentos,     
        "link": curso.link,
        "tipo_horario": curso.tipo_oferta,  
    })

    return render(request, "formularios/formulario-formato.html", {
        "form": form,
        "usuario": curso.usuario,
        "modo_detalle": True,
    })

#Reportes
def reportes(request):
    cursos = Curso.objects.annotate(
        inscripciones_count=Count("aprendices")
    ).order_by('-fecha_inicio')
    return render(request, 'reportes.html', {'cursos': cursos})

def generate_reports(request, curso_id):
    course = Curso.objects.get(id=curso_id)
    aspirantes = Aspirante.objects.filter(curso=course)

    # ---------- REPORTE 1 ----------
    aspirantes_data1 = []
    for asp in aspirantes:
        partes = asp.nombre.strip().split(" ", 1)
        nombres = partes[0] if len(partes) > 0 else ""
        apellidos = partes[1] if len(partes) > 1 else ""

        aspirantes_data1.append([
            asp.tipo_documento.nombre if asp.tipo_documento else "",
            asp.documento,
            nombres,
            apellidos,
            asp.correo,
            asp.telefono if asp.telefono else "",
            asp.poblacion.nombre if asp.poblacion else "",
            course.programa.nombre if course.programa else "",
        ])

    df1 = pd.DataFrame(aspirantes_data1, columns=[
        "Tipo Documento", "Documento", "Nombres", "Apellidos",
        "Correo", "Teléfono", "Población", "Programa"
    ])

    # ---------- REPORTE 2 ----------
    plantilla_path = os.path.join(settings.BASE_DIR, "curso", "templates", "docs", "Masivo 2024 MOISO.xlsm")
    wb = load_workbook(plantilla_path)
    ws = wb.active
    fila = 3
    for asp in aspirantes:
        ws[f"B{fila}"] = asp.tipo_documento.nombre if asp.tipo_documento else ""
        ws[f"C{fila}"] = asp.documento
        ws[f"E{fila}"] = asp.poblacion.nombre if asp.poblacion else ""
        fila += 1

    # 📂 Carpeta del curso (donde están los PDFs también)
    carpeta_curso = os.path.join(settings.BASE_DIR, "curso", "templates", "docs", str(curso_id))
    os.makedirs(carpeta_curso, exist_ok=True)

    # Guardar Reporte 1
    path_reporte1 = os.path.join(carpeta_curso, f"Reporte_Completo_curso_{curso_id}.xlsx")
    df1.to_excel(path_reporte1, index=False, engine="openpyxl")

    # Guardar Reporte 2
    path_reporte2 = os.path.join(carpeta_curso, f"Reporte_Poblacion_curso_{curso_id}.xlsx")
    wb.save(path_reporte2)

    # Crear ZIP con reportes y PDFs
    zip_path = os.path.join(carpeta_curso, f"Reportes_curso_{curso_id}.zip")
    with zipfile.ZipFile(zip_path, "w") as zipf:
        # Añadir reportes Excel (siempre se generan, aunque no haya aspirantes)
        zipf.write(path_reporte1, os.path.basename(path_reporte1))
        zipf.write(path_reporte2, os.path.basename(path_reporte2))

        # Añadir consolidado si existe
        consolidado = os.path.join(carpeta_curso, "aspirantes_consolidado.pdf")
        if os.path.exists(consolidado):
            zipf.write(consolidado, "aspirantes_consolidado.pdf")

        # Añadir PDFs individuales de aspirantes
        carpeta_aspirantes = os.path.join(carpeta_curso, "aspirantes")
        if os.path.exists(carpeta_aspirantes):
            for archivo in os.listdir(carpeta_aspirantes):
                if archivo.lower().endswith(".pdf"):
                    ruta_pdf = os.path.join(carpeta_aspirantes, archivo)
                    zipf.write(ruta_pdf, os.path.join("aspirantes", archivo))

    # Descargar el ZIP
    return FileResponse(open(zip_path, "rb"), as_attachment=True, filename=f"Reportes_curso_{curso_id}.zip")

# Generar curso y documento
def generar_curso(request, tipo):
    usuario = request.user  # Usuario autenticado

    if request.method == "POST":
        form = CursoForm(request.POST, request.FILES, usuario=usuario)
        if form.is_valid():
            try:
                # ✅ Crear curso en BD
                programa = form.cleaned_data["nombreprograma"]
                fecha_inicio = form.cleaned_data["fechainicio"]
                fecha_fin = form.cleaned_data["fechafin"]

                curso = Curso.objects.create(
                    programa=programa,
                    usuario=usuario,
                    fecha_inicio=fecha_inicio,
                    fecha_fin=fecha_fin,
                    tipo_oferta=tipo,
                )

                # ✅ Preparar horario
                tipo_horario = request.POST.get("tipo_horario", "general")
                if tipo_horario == "general":
                    hora_inicio = form.cleaned_data.get("horario_inicio")
                    hora_fin = form.cleaned_data.get("horario_fin")
                    horario_texto = (
                        f"{hora_inicio.strftime('%H:%M')} - {hora_fin.strftime('%H:%M')}"
                        if hora_inicio and hora_fin else ""
                    )
                elif tipo_horario == "individual":
                    dias = form.cleaned_data.get("dias", [])
                    horarios_individuales = []
                    for dia in dias:
                        h_ini = request.POST.get(f"hora_inicio_{dia}")
                        h_fin = request.POST.get(f"hora_fin_{dia}")
                        if h_ini and h_fin:
                            horarios_individuales.append(f"{dia}: {h_ini} - {h_fin}")
                    horario_texto = ", ".join(horarios_individuales)
                else:
                    horario_texto = ""

                # ✅ Contexto para plantilla
                contexto = form.cleaned_data
                contexto.update({
                    "nombre": f"{usuario.first_name} {usuario.last_name}".strip(),
                    "tipodoc": usuario.tipo_documento.nombre if usuario.tipo_documento else "",
                    "numerodoc": usuario.documento,
                    "correo": usuario.email,
                    "curso_id": curso.id,
                    "horario": horario_texto,
                })

                # ✅ Generar documento Word
                ruta_base = os.path.join(settings.BASE_DIR, "curso", "templates", "docs")
                ruta_plantilla = os.path.join(ruta_base, "plantilla-curso.docx")
                if not os.path.exists(ruta_plantilla):
                    raise FileNotFoundError(f"No se encontró la plantilla en: {ruta_plantilla}")

                ruta_curso = os.path.join(ruta_base, str(curso.id))
                os.makedirs(ruta_curso, exist_ok=True)

                doc = DocxTemplate(ruta_plantilla)  # 👈 se crea una sola vez

                # ✅ Agregar firma como imagen si existe
                if usuario.firma_digital:
                    contexto["firma"] = InlineImage(
                        doc,
                        usuario.firma_digital.path,
                        width=Mm(40)
                    )
                else:
                    contexto["firma"] = ""

                # Renderizar y guardar
                doc.render(contexto)
                nombre_docx = f"curso_{curso.id}.docx"
                ruta_docx = os.path.join(ruta_curso, nombre_docx)
                doc.save(ruta_docx)

                # ✅ Guardar carta si se adjunta
                carta_file = request.FILES.get("carta_empresa")
                if carta_file:
                    nombre_carta = f"carta_{curso.id}{os.path.splitext(carta_file.name)[1]}"
                    ruta_carta = os.path.join(ruta_curso, nombre_carta)
                    with open(ruta_carta, "wb+") as destino:
                        for chunk in carta_file.chunks():
                            destino.write(chunk)
                    curso.carta = f"docs/{curso.id}/{nombre_carta}"

                # Guardar ruta del documento en el curso
                curso.caracterizacion = f"docs/{curso.id}/{nombre_docx}"
                curso.save(update_fields=["caracterizacion", "carta"])

                # ✅ Generar URL de aspirantes
                url_aspirantes = request.build_absolute_uri(
                    reverse("registrar_aspirante", args=[curso.id])
                )

                # ✅ Guardar en sesión
                request.session["curso_id"] = curso.id
                request.session["url_aspirantes"] = url_aspirantes
                request.session["ruta_docx"] = curso.caracterizacion

                return redirect("curso_generado")

            except Exception as e:
                return HttpResponse(f"Error generando el documento: {e}", status=500)

    else:
        form = CursoForm(usuario=usuario)

    return render(request, "formularios/formulario-formato.html", {"form": form, "tipo": tipo, "modo_detalle": False})

def tipo_oferta(request):
    return render(request, "tipo_oferta.html")

def filtrar_programas(request):
    area_id = request.GET.get("area")
    duracion = request.GET.get("duracion")

    programas = Programa.objects.all()

    if area_id:
        programas = programas.filter(area_id=area_id)
    if duracion:
        programas = programas.filter(duracion=duracion)

    data = [
        {"id": p.id, "nombre": p.nombre, "codigo": p.codigo, "version": p.version, "duracion": p.duracion}
        for p in programas
    ]
    return JsonResponse({"programas": data})

# Obtener datos del programa
def get_programa(request, programa_id):
    try:
        programa = Programa.objects.get(id=programa_id)
        data = {
            "codigo": programa.codigo,
            "version": programa.version,
            "duracion": programa.duracion,
        }
        return JsonResponse(data)
    except Programa.DoesNotExist:
        return JsonResponse({"error": "Programa no encontrado"}, status=404)
    
# Obtener datos de departamento y municipio
def get_municipios(request, departamento_id):
    municipios = Municipio.objects.filter(departamento_id=departamento_id).values("id", "nombre")
    return JsonResponse(list(municipios), safe=False)

# Página que muestra el ID del curso y la URL para registrar aspirantes
def curso_generado(request):
    curso_id = request.session.get("curso_id")
    url_aspirantes = request.session.get("url_aspirantes")

    if not curso_id:
        return redirect("generar_curso")

    curso = Curso.objects.select_related("programa").get(id=curso_id)

    # 👇 Guardar la URL en el modelo si existe
    if url_aspirantes:
        curso.link = url_aspirantes
        curso.save(update_fields=["link"])

    return render(request, "formularios/curso_generado.html", {
        "curso": curso,
        "url_aspirantes": url_aspirantes,
    })

# Registrar aspirante a un curso
def registrar_aspirante(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)

    if request.method == "POST":
        form = AspiranteForm(request.POST, request.FILES)
        if form.is_valid():
            aspirante = form.save(commit=False)
            aspirante.curso = curso  

            if "archivo_documento" in request.FILES:
                archivo = request.FILES["archivo_documento"]

                # 📂 Carpeta del curso
                carpeta_curso = os.path.join(
                    settings.BASE_DIR, "curso", "templates", "docs", str(curso.id), "aspirantes"
                )
                os.makedirs(carpeta_curso, exist_ok=True)
                
                url_pdf_global = os.path.join(
                    settings.BASE_DIR, "curso", "templates", "docs", str(curso.id)
                )
                os.makedirs(url_pdf_global, exist_ok=True)

                fs = FileSystemStorage(location=carpeta_curso)
                nombre_archivo = fs.save(archivo.name, archivo)

                # Guardar la ruta relativa en el modelo
                aspirante.archivo_documento.name = f"docs/{curso.id}/aspirantes/{nombre_archivo}"

                # 📌 Consolidado por curso
                pdf_padre_path = os.path.join(url_pdf_global, "aspirantes_consolidado.pdf")
                pdf_nuevo_path = os.path.join(carpeta_curso, nombre_archivo)

                merger = PdfMerger()

                # Si ya existe un consolidado, lo cargamos primero
                if os.path.exists(pdf_padre_path):
                    merger.append(pdf_padre_path)

                # Agregamos el nuevo PDF al final
                merger.append(pdf_nuevo_path)

                # Guardamos en archivo temporal para evitar corrupción
                with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_file:
                    merger.write(tmp_file.name)
                    merger.close()

                # Reemplazamos el consolidado con el archivo temporal
                shutil.move(tmp_file.name, pdf_padre_path)

            aspirante.save()

            messages.success(
                request,
                f"Aspirante {aspirante.nombre} registrado con éxito en el curso {curso.programa.nombre}."
            )
            return redirect("registrar_aspirante", curso_id=curso.id)
        else:
            messages.error(request, "Por favor corrige los errores en el formulario.")
    else:
        form = AspiranteForm()

    return render(request, "formularios/registrar_aspirante.html", {
        "curso": curso,
        "form": form
    })

# OCR para extraer datos del documento
# Ruta al ejecutable de Tesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

@csrf_exempt
def ocr_aspirante(request):
    if request.method == "POST" and request.FILES.get("archivo_documento"):
        archivo = request.FILES["archivo_documento"]

        # Guardar archivo temporal
        temp_path = os.path.join(settings.MEDIA_ROOT, archivo.name)
        with open(temp_path, "wb+") as destino:
            for chunk in archivo.chunks():
                destino.write(chunk)

        texto = ""

        try:
            if archivo.name.lower().endswith(".pdf"):
                # Convertir PDF a imágenes (con poppler_path)
                pages = convert_from_path(
                    temp_path,
                    dpi=300,
                    poppler_path=r"C:\Users\SENASoft\Downloads\Release-25.07.0-0\poppler-25.07.0\Library\bin"
                )
                for page in pages:
                    texto += pytesseract.image_to_string(page, lang="spa")
                    print(texto)
            else:
                # Imagen directa
                img = Image.open(temp_path)
                texto = pytesseract.image_to_string(img, lang="spa")

            # Regex para capturar datos
            # Normalizar todo el texto para buscar números
            texto_numeros = re.sub(r"[^0-9]", "", texto)  # dejar solo dígitos

            # Buscar el número con search
            match = re.search(r"\d{10}", texto_numeros)
            numero_doc = match.group() if match else ""   # aquí group está bien
            # Intento 1: buscar con etiquetas
            # search(r"NOMBRES?\s*[\r\n]+([A-ZÁÉÍÓÚÑ\s]+)"
            # search(r"APELLIDOS?\s*[\r\n]+([A-ZÁÉÍÓÚÑ\s]+)"    
            apellidos = re.search(r"NOMBRES?\s*[\r\n]+([A-ZÁÉÍÓÚÑ\s]+)", texto, re.IGNORECASE)
            nombres = re.search(r"APELLIDOS?\s*[\r\n]+([A-ZÁÉÍÓÚÑ\s]+)", texto, re.IGNORECASE)

            print(numero_doc, apellidos, nombres)

            # Fallback: bloques de mayúsculas (por si OCR falla con etiquetas)
            if not (apellidos and nombres):
                mayusculas = re.findall(r"[A-ZÁÉÍÓÚÑ]{2,}(?:\s+[A-ZÁÉÍÓÚÑ]{2,})+", texto)
                if len(mayusculas) >= 2:
                    apellidos = mayusculas[0]
                    nombres = mayusculas[1]
                else:
                    apellidos = ""
                    nombres = ""

            # Construcción del JSON (soporta tanto regex normal como fallback string)
            data = {
                "documento": numero_doc if numero_doc else "",
                "apellidos": apellidos.group(1).strip() if hasattr(apellidos, "group") else apellidos,
                "nombres": nombres.group(1).strip() if hasattr(nombres, "group") else nombres,
                "texto_completo": texto
            }
                        
            return JsonResponse(data)

        except Exception as e:
            import traceback
            traceback.print_exc()  # imprime el error completo en consola
            return JsonResponse(
                {"error": f"Ocurrió un problema procesando el archivo: {str(e)}"},
                status=500
            )

        finally:
            # Eliminar archivo temporal
            if os.path.exists(temp_path):
                os.remove(temp_path)

    return JsonResponse({"error": "Archivo no válido"}, status=400)

# Descargar documento generado
def descargar_curso(request, curso_id):
    try:
        curso = Curso.objects.get(id=curso_id)
        ruta_docx = os.path.join(settings.BASE_DIR, "curso", "templates", curso.caracterizacion)

        if not os.path.exists(ruta_docx):
            raise Http404("Archivo no encontrado")

        return FileResponse(
            open(ruta_docx, "rb"),
            as_attachment=True,
            filename=f"curso_{curso.id}.docx"
        )

    except Curso.DoesNotExist:
        raise Http404("Curso no existe")

def inicioSesion(request):
    if request.method == 'POST':
        form = InicioSesionForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data['email']
            password = form.cleaned_data['password']
            # Autenticar al usuario con el email y password
            usuario = authenticate(request, email=email, password=password)
            if usuario is not None:
                login(request, usuario)
                messages.success(request, f"Inicio de sesión exitoso para {email}.")
                
                # Redirigir según el rol del usuario
                if usuario.rol.nombre == "INSTRUCTOR":
                    return redirect('dashboard')  # Dashboard para instructores
                elif usuario.rol.nombre == "FUNCIONARIO":
                    return redirect('dashboard')  # Panel de coordinación para funcionarios
                elif usuario.rol.nombre == "Admin" or usuario.is_superuser:
                    return redirect('viewUsuarios')  # Vista de administración para admins
                else:
                    # Redirección por defecto para otros roles no especificados
                    return redirect('dashboard')
            else:
                messages.error(request, "Credenciales inválidas.")
    else:
        form = InicioSesionForm()
    return render(request, 'layout/inicioSesion.html', {'form': form})

def log_out(request):
    logout(request)  # elimina la sesión del usuario
    messages.success(request, "Has cerrado sesión correctamente.")
    return redirect('inicioSesion')  

@login_required
def usuario(request):
    form = UsuarioCreateForm()
    return render(request, 'layout/usuario_form.html', {'form': form })

@login_required
def createUsuario(request):
    if request.method == 'POST':
        form = UsuarioCreateForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario creado correctamente.")
            return redirect('viewUsuarios')
    else:
        form = UsuarioCreateForm()
    return render(request, 'layout/usuario_form.html', {"form": form})

@login_required
def viewUsuarios(request):
    usuarios= Usuario.objects.all()
    return render(request, 'layout/viewUsuarios.html', {'usuarios': usuarios})

@login_required
def editUsuario(request, id):
    usuario = get_object_or_404(Usuario, id=id)
    if request.method == 'POST':
        form = UsuarioEditForm(request.POST, request.FILES, instance=usuario)
        if form.is_valid():
            form.save()
            messages.success(request, f"El usuario {usuario.email} ha sido actualizado.")
            return redirect('viewUsuarios')
    else:
        form = UsuarioEditForm(instance=usuario)
    return render(request, 'layout/usuario_form.html', {"form": form})

@login_required
def toggle_usuario(request, id):
    usuario = get_object_or_404(Usuario, id=id)
    usuario.is_active = not usuario.is_active  # Cambiar True <-> False
    usuario.save()
    
    estado = "activado" if usuario.is_active else "desactivado"
    messages.success(request, f"El usuario {usuario.email} ha sido {estado}.")
    return redirect("viewUsuarios")  # redirige a tu lista de usuarios

@login_required
def subir_firma(request):
    if request.method == 'POST':
        archivo = request.FILES.get('firma')
        
        if archivo:
            # Validar tipo de archivo
            if archivo.content_type not in ['image/png', 'image/jpeg']:
                messages.error(request, "Formato no válido. Use PNG o JPG")
                return redirect('subir_firma')
            
            # Validar dimensiones de la imagen
            try:
                width, height = get_image_dimensions(archivo)
                if width > 500 or height > 200:
                    messages.error(request, "La imagen no debe exceder 500x200 píxeles")
                    return redirect('subir_firma')
            except Exception:
                messages.error(request, "Error al procesar la imagen")
                return redirect('subir_firma')
            
            # Guardar la firma
            usuario = request.user
            primera_vez = not bool(usuario.firma_digital)
            
            usuario.firma_digital = archivo
            usuario.save()
            
            if primera_vez:
                messages.success(request, "Firma digital registrada correctamente")
            else:
                messages.success(request, "Firma digital actualizada correctamente")
            
            return redirect('dashboard')
        else:
            messages.error(request, "Debe seleccionar un archivo")
    
    return render(request, 'subir_firma.html')